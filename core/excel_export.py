"""Exportação para Excel com formatação brasileira."""
import pandas as pd
from datetime import datetime
import io
from typing import Dict, Optional
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils.dataframe import dataframe_to_rows

def export_to_excel(df: pd.DataFrame, stats: Optional[Dict] = None) -> bytes:
    """
    Exporta dados para Excel com múltiplas abas, gráficos embarcados e formatação avançada.
    
    Args:
        df: DataFrame com os dados
        stats: Dicionário com estatísticas descritivas
    
    Returns:
        Bytes do arquivo Excel
    """
    output = io.BytesIO()
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create styles
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=None, right=None, top=None, bottom=None)
    
    # Sheet 1: Dados
    ws_data = wb.create_sheet("Dados")
    
    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    # Format headers
    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Apply conditional formatting
    _apply_conditional_formatting(ws_data, df)
    
    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Estatísticas
    if stats:
        ws_stats = wb.create_sheet("Estatísticas")
        stats_df = pd.DataFrame([stats])
        
        for r in dataframe_to_rows(stats_df, index=False, header=True):
            ws_stats.append(r)
        
        # Format headers
        for cell in ws_stats[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    # Sheet 3: Pivot Tables
    ws_pivot = wb.create_sheet("Tabelas Dinâmicas")
    _create_pivot_tables(ws_pivot, df)
    
    # Sheet 4: Gráficos
    ws_charts = wb.create_sheet("Gráficos")
    _add_embedded_charts(ws_charts, df)
    
    # Sheet 5: Resumo
    ws_summary = wb.create_sheet("Resumo")
    _create_summary_sheet(ws_summary, df, stats)
    
    # Save to bytes
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def _apply_conditional_formatting(worksheet, df):
    """Apply conditional formatting to the data worksheet."""
    
    # Color scale for price columns
    price_columns = ['PREÇO POR KG', 'VALOR', 'ARROBA GORDO', 'ARROBA MAGRO']
    for col_name in price_columns:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1  # Excel is 1-indexed
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            
            # Color scale (red to green)
            color_scale = ColorScaleRule(
                start_type="min", start_color="FF0000",
                mid_type="percentile", mid_value=50, mid_color="FFFF00",
                end_type="max", end_color="00FF00"
            )
            
            # Apply to data range (skip header)
            data_range = f"{col_letter}2:{col_letter}{len(df)+1}"
            worksheet.conditional_formatting.add(data_range, color_scale)
    
    # Data bars for weight column
    if 'PESO (KG)' in df.columns:
        col_idx = df.columns.get_loc('PESO (KG)') + 1
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        
        data_bar = DataBarRule(
            start_type="min", start_value=0,
            end_type="max", end_value=None,
            color="4CAF50"
        )
        
        data_range = f"{col_letter}2:{col_letter}{len(df)+1}"
        worksheet.conditional_formatting.add(data_range, data_bar)

def _create_pivot_tables(worksheet, df):
    """Create pivot tables for data analysis."""
    
    # Add title
    worksheet['A1'] = "Tabelas Dinâmicas"
    worksheet['A1'].font = Font(size=16, bold=True)
    
    # Pivot 1: Estado x Raça com preços médios
    if 'ESTADO' in df.columns and 'RAÇA' in df.columns and 'PREÇO POR KG' in df.columns:
        worksheet['A3'] = "Preço Médio por Estado e Raça"
        worksheet['A3'].font = Font(size=14, bold=True)
        
        # Create pivot data
        pivot_data = df.groupby(['ESTADO', 'RAÇA'])['PREÇO POR KG'].mean().reset_index()
        pivot_data = pivot_data.pivot(index='ESTADO', columns='RAÇA', values='PREÇO POR KG').fillna(0)
        
        # Add pivot table
        start_row = 5
        for r in dataframe_to_rows(pivot_data, index=True, header=True):
            worksheet.append(r)
        
        # Format pivot table
        for cell in worksheet[start_row]:
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.font = Font(bold=True)
    
    # Pivot 2: Ano x Trimestre com totais
    if 'ANO' in df.columns and 'TRIMESTRE' in df.columns:
        worksheet['A15'] = "Total de Registros por Ano e Trimestre"
        worksheet['A15'].font = Font(size=14, bold=True)
        
        # Create pivot data
        pivot_data = df.groupby(['ANO', 'TRIMESTRE']).size().reset_index(name='Total')
        pivot_data = pivot_data.pivot(index='ANO', columns='TRIMESTRE', values='Total').fillna(0)
        
        # Add pivot table
        start_row = 17
        for r in dataframe_to_rows(pivot_data, index=True, header=True):
            worksheet.append(r)
        
        # Format pivot table
        for cell in worksheet[start_row]:
            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            cell.font = Font(bold=True)

def _add_embedded_charts(worksheet, df):
    """Add embedded charts to the worksheet."""
    
    # Add title
    worksheet['A1'] = "Gráficos de Análise"
    worksheet['A1'].font = Font(size=16, bold=True)
    
    # Chart 1: Top 10 Estados (Bar Chart)
    if 'ESTADO' in df.columns:
        state_counts = df['ESTADO'].value_counts().head(10)
        
        # Add data
        worksheet['A3'] = "Estado"
        worksheet['B3'] = "Quantidade"
        
        for i, (state, count) in enumerate(state_counts.items(), start=4):
            worksheet[f'A{i}'] = state
            worksheet[f'B{i}'] = count
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Top 10 Estados por Quantidade"
        chart.x_axis.title = "Estado"
        chart.y_axis.title = "Quantidade"
        
        data = Reference(worksheet, min_col=2, min_row=3, max_row=3+len(state_counts))
        cats = Reference(worksheet, min_col=1, min_row=4, max_row=4+len(state_counts))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        worksheet.add_chart(chart, "D3")
    
    # Chart 2: Distribuição de Raças (Pie Chart)
    if 'RAÇA' in df.columns:
        breed_counts = df['RAÇA'].value_counts().head(8)
        
        # Add data (offset to avoid overlap)
        start_row = 20
        worksheet[f'A{start_row}'] = "Raça"
        worksheet[f'B{start_row}'] = "Quantidade"
        
        for i, (breed, count) in enumerate(breed_counts.items(), start=start_row+1):
            worksheet[f'A{i}'] = breed
            worksheet[f'B{i}'] = count
        
        # Create pie chart
        pie_chart = PieChart()
        pie_chart.title = "Distribuição por Raça"
        
        data = Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row+len(breed_counts))
        labels = Reference(worksheet, min_col=1, min_row=start_row+1, max_row=start_row+1+len(breed_counts))
        
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        
        worksheet.add_chart(pie_chart, "D20")
    
    # Chart 3: Tendência de Preços (Line Chart)
    if 'ANO' in df.columns and 'PREÇO POR KG' in df.columns:
        price_trend = df.groupby('ANO')['PREÇO POR KG'].mean()
        
        # Add data (offset to avoid overlap)
        start_row = 35
        worksheet[f'A{start_row}'] = "Ano"
        worksheet[f'B{start_row}'] = "Preço Médio"
        
        for i, (year, price) in enumerate(price_trend.items(), start=start_row+1):
            worksheet[f'A{i}'] = year
            worksheet[f'B{i}'] = price
        
        # Create line chart
        line_chart = LineChart()
        line_chart.title = "Tendência de Preços por Ano"
        line_chart.x_axis.title = "Ano"
        line_chart.y_axis.title = "Preço Médio (R$/kg)"
        
        data = Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row+len(price_trend))
        cats = Reference(worksheet, min_col=1, min_row=start_row+1, max_row=start_row+1+len(price_trend))
        
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(cats)
        
        worksheet.add_chart(line_chart, "D35")

def _create_summary_sheet(worksheet, df, stats):
    """Create a summary sheet with key metrics."""
    
    # Add title
    worksheet['A1'] = "Resumo Executivo"
    worksheet['A1'].font = Font(size=16, bold=True)
    
    # Basic info
    summary_data = [
        ["Métrica", "Valor"],
        ["Total de Registros", len(df)],
        ["Total de Colunas", len(df.columns)],
        ["Data de Geração", datetime.now().strftime('%d/%m/%Y %H:%M')],
        ["", ""],  # Empty row
    ]
    
    # Add price statistics if available
    if 'PREÇO POR KG' in df.columns:
        price_col = df['PREÇO POR KG'].dropna()
        if not price_col.empty:
            summary_data.extend([
                ["PREÇO POR KG", ""],
                ["  Média", f"R$ {price_col.mean():.2f}"],
                ["  Mediana", f"R$ {price_col.median():.2f}"],
                ["  Mínimo", f"R$ {price_col.min():.2f}"],
                ["  Máximo", f"R$ {price_col.max():.2f}"],
                ["", ""],  # Empty row
            ])
    
    # Add weight statistics if available
    if 'PESO (KG)' in df.columns:
        weight_col = df['PESO (KG)'].dropna()
        if not weight_col.empty:
            summary_data.extend([
                ["PESO (KG)", ""],
                ["  Média", f"{weight_col.mean():.1f} kg"],
                ["  Mediana", f"{weight_col.median():.1f} kg"],
                ["  Mínimo", f"{weight_col.min():.1f} kg"],
                ["  Máximo", f"{weight_col.max():.1f} kg"],
                ["", ""],  # Empty row
            ])
    
    # Add categorical summaries
    categorical_cols = df.select_dtypes(include=['object']).columns[:3]
    for col in categorical_cols:
        top_values = df[col].value_counts().head(3)
        summary_data.append([f"Top {col}", ""])
        for value, count in top_values.items():
            summary_data.append([f"  {value}", f"{count} ({count/len(df)*100:.1f}%)"])
        summary_data.append(["", ""])  # Empty row
    
    # Add data to worksheet
    for row_data in summary_data:
        worksheet.append(row_data)
    
    # Format headers
    worksheet['A2'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    worksheet['A2'].font = Font(bold=True, color="FFFFFF")
    worksheet['B2'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    worksheet['B2'].font = Font(bold=True, color="FFFFFF")
    
    # Auto-adjust column widths
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 20

